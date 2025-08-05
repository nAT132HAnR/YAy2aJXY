# 代码生成时间: 2025-08-06 03:54:18
import os
import datetime
from sanic import Sanic
from sanic.response import json, file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Initialize the Sanic application
app = Sanic('ExcelGeneratorService')

# Define the route for generating an Excel file
@app.route('/create_excel', methods=['GET', 'POST'])
async def create_excel(request):
    # Extract data from request
    data = request.json
    if not data or 'columns' not in data or 'rows' not in data:
        return json({'error': 'Missing data or columns/rows'}, status=400)

    try:
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'GeneratedData'

        # Set column titles
        for col, value in enumerate(data['columns'], start=1):
            ws.cell(row=1, column=col).value = value
            # Style the header
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

        # Fill the rows with data
        for row_index, row_data in enumerate(data['rows'], start=2):
            for col_index, value in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_index).value = value

        # Save the workbook to a file
        file_name = 'generated_excel_{}.xlsx'.format(datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
        wb.save(file_name)

        # Return the file as a response
        return file(file_name, mime_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return json({'error': 'Failed to generate Excel file', 'details': str(e)}, status=500)

# Run the application
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=True)

"""
Excel Generator Service
===================

This service allows the generation of Excel files dynamically based on the provided data.

Endpoints:
- /create_excel: Accepts data in JSON format and returns an Excel file.

Request JSON structure:
{
  "columns": ["Column1", "Column2", ...],
  "rows": [
    ["Value11", "Value12", ...],
    ["Value21", "Value22", ...],
    ...
  ]
}

Error Handling:
- If the data is missing or incorrectly formatted, the service returns a 400 status code with an error message.
- Any failures in generating the Excel file result in a 500 status code with an error message.

Dependencies:
- openpyxl: For creating and saving Excel files.
- pandas: For data manipulation (optional and not used in this snippet).

Usage:
To use this service, send a POST request to /create_excel with the JSON data in the body.
The service will respond with an Excel file attached to the response.

"""